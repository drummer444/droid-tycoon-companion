"""
Regenerates excel/DroidTycoonCompanion.xlsx from database/droids.json and
database/rebirths.json.

Run from the repo root:
    python scripts/build_excel.py

This produces a *fresh template* — droid list, income values, and rebirth
requirements are rebuilt from the JSON database, but the Quick Entry sheet
is always reset to blank/unowned. Personal ownership progress is NOT stored
in the JSON database (on purpose — that's your own save data, not shared
game data), so it is never overwritten or synced by this script. Keep your
own working copy of the .xlsx separate from the one this script outputs if
you don't want to lose your Quick Entry progress when the database updates.
"""
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule

ROOT = Path(__file__).resolve().parent.parent
DROIDS_PATH = ROOT / "database" / "droids.json"
REBIRTHS_PATH = ROOT / "database" / "rebirths.json"
OUT_PATH = ROOT / "excel" / "DroidTycoonCompanion.xlsx"

FONT = "Arial"
HEADER_FILL = PatternFill("solid", start_color="1B1F3B", end_color="1B1F3B")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
SUBHEADER_FILL = PatternFill("solid", start_color="D9DCE6", end_color="D9DCE6")
INPUT_FILL = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")
BLUE = Font(name=FONT, color="0000FF")
BLACK = Font(name=FONT, color="000000")
BOLD = Font(name=FONT, bold=True)
TITLE_FONT = Font(name=FONT, bold=True, size=16, color="1B1F3B")
thin = Side(style="thin", color="B7B7B7")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_header(ws, row, col_start, col_end):
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def autosize(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def main():
    droids_data = json.loads(DROIDS_PATH.read_text())["droids"]
    rebirths_data = json.loads(REBIRTHS_PATH.read_text())["rebirths"]

    wb = Workbook()

    # ---------------- Lists ----------------
    lists = wb.active
    lists.title = "Lists"
    tiers = [("None", 0), ("Base", 1), ("Gold", 2), ("Diamond", 3), ("Rainbow", 4), ("Beskar", 5)]
    lists["A1"] = "Tier"
    lists["B1"] = "Rank"
    for i, (t, r) in enumerate(tiers, start=2):
        lists[f"A{i}"] = t
        lists[f"B{i}"] = r
    lists["D1"] = "YesNo"
    lists["D2"] = "Yes"
    lists["D3"] = "No"
    lists["F1"] = "OwnedTierChoices"
    for i, t in enumerate(["Base", "Gold", "Diamond", "Rainbow", "Beskar"], start=2):
        lists[f"F{i}"] = t

    # ---------------- Droid Database ----------------
    db = wb.create_sheet("Droid Database")
    headers = ["Name", "Type", "Rarity", "Base Income/s", "Gold Income/s", "Diamond Income/s",
               "Rainbow Income/s", "Owned Tier (auto)", "Owned Rank", "Current Income/s", "Notes"]
    for i, h in enumerate(headers, start=1):
        db.cell(row=1, column=i, value=h)
    style_header(db, 1, 1, len(headers))
    db.freeze_panes = "A2"

    start_row = 2
    for i, d in enumerate(droids_data, start=start_row):
        inc = d["income"]
        db.cell(row=i, column=1, value=d["name"]).font = BLACK
        db.cell(row=i, column=2, value=d["type"])
        db.cell(row=i, column=3, value=d["rarity"])
        db.cell(row=i, column=4, value=inc.get("base") if inc.get("base") is not None else "TBC").font = BLUE
        db.cell(row=i, column=5, value=inc.get("gold") if inc.get("gold") is not None else "TBC").font = BLUE
        db.cell(row=i, column=6, value=inc.get("diamond") if inc.get("diamond") is not None else "TBC").font = BLUE
        db.cell(row=i, column=7, value=inc.get("rainbow") if inc.get("rainbow") is not None else "TBC").font = BLUE
        db.cell(row=i, column=8,
                value=(f'=IFERROR(IF(INDEX(\'Quick Entry\'!$B$2:$B$60,MATCH(A{i},\'Quick Entry\'!$A$2:$A$60,0))="Yes",'
                       f'INDEX(\'Quick Entry\'!$C$2:$C$60,MATCH(A{i},\'Quick Entry\'!$A$2:$A$60,0)),"None"),"None")'))
        db.cell(row=i, column=9, value=f"=IFERROR(VLOOKUP(H{i},Lists!$A$2:$B$7,2,FALSE),0)").font = BLACK
        db.cell(row=i, column=10,
                value=f'=IF(H{i}="None",0,IF(H{i}="Base",D{i},IF(H{i}="Gold",E{i},'
                      f'IF(H{i}="Diamond",F{i},IF(H{i}="Rainbow",G{i},"N/A (TBC)")))))').font = BLACK
        db.cell(row=i, column=11, value=d.get("notes") or "")
        for c in range(1, 12):
            db.cell(row=i, column=c).border = BORDER

    last_droid_row = start_row + len(droids_data) - 1
    autosize(db, {"A": 18, "B": 12, "C": 12, "D": 14, "E": 14, "F": 16, "G": 16,
                  "H": 12, "I": 10, "J": 16, "K": 55})

    # ---------------- Quick Entry ----------------
    qe = wb.create_sheet("Quick Entry")
    qe["A1"] = "Add your droids here"
    qe["A1"].font = TITLE_FONT
    qe["A2"] = ("Pick a droid from the dropdown, mark Owned = Yes, then pick the tier you have. "
                "That's it — the Droid Database and Rebirth Tracker update automatically.")
    qe["A2"].font = Font(name=FONT, italic=True, size=10, color="666666")
    qe.merge_cells("A2:E2")

    for i, h in enumerate(["Droid", "Owned?", "Tier"], start=1):
        qe.cell(row=4, column=i, value=h)
    style_header(qe, 4, 1, 3)

    QE_START, QE_END = 5, 63
    for r in range(QE_START, QE_END + 1):
        qe.cell(row=r, column=2, value="No")
        for c in range(1, 4):
            qe.cell(row=r, column=c).border = BORDER

    dv_droid = DataValidation(type="list", formula1=f"='Droid Database'!$A${start_row}:$A${last_droid_row}", allow_blank=True)
    qe.add_data_validation(dv_droid)
    dv_droid.add(f"A{QE_START}:A{QE_END}")

    dv_yesno = DataValidation(type="list", formula1="=Lists!$D$2:$D$3", allow_blank=True)
    qe.add_data_validation(dv_yesno)
    dv_yesno.add(f"B{QE_START}:B{QE_END}")

    dv_tier = DataValidation(type="list", formula1="=Lists!$F$2:$F$6", allow_blank=True)
    qe.add_data_validation(dv_tier)
    dv_tier.add(f"C{QE_START}:C{QE_END}")

    green_fill_qe = PatternFill("solid", start_color="C6EFCE", end_color="C6EFCE")
    qe.conditional_formatting.add(f"A{QE_START}:C{QE_END}",
                                   FormulaRule(formula=[f"$B{QE_START}=\"Yes\""], fill=green_fill_qe))
    autosize(qe, {"A": 20, "B": 12, "C": 14})

    # ---------------- Upgrade Costs ----------------
    uc = wb.create_sheet("Upgrade Costs")
    uc["A1"] = "Upgrade Chip Costs by Rarity"
    uc["A1"].font = TITLE_FONT
    for i, h in enumerate(["Rarity", "Chips: Base -> Gold", "Chips: Gold -> Diamond", "Chips: Diamond -> Rainbow"], start=1):
        uc.cell(row=3, column=i, value=h)
    style_header(uc, 3, 1, 4)
    costs = [("Common", 5, 10, 15), ("Rare", 30, 50, 75), ("Epic", 120, 180, 240),
             ("Legendary", 400, 1200, 4000), ("Mythic", "TBC", "TBC", "TBC")]
    for i, row_data in enumerate(costs, start=4):
        for j, val in enumerate(row_data, start=1):
            cell = uc.cell(row=i, column=j, value=val)
            cell.font = BLUE if j > 1 else BLACK
            cell.border = BORDER
    uc["A10"] = "Source: Insider Gaming, Droidex guide. Cross-check in-game before heavy chip investment."
    uc["A10"].font = Font(name=FONT, italic=True, size=9, color="666666")
    autosize(uc, {"A": 16, "B": 20, "C": 22, "D": 24})

    # ---------------- Rebirth Tracker ----------------
    rt = wb.create_sheet("Rebirth Tracker")
    rt_headers = ["Rebirth", "Credits Needed", "Credits Ready?",
                  "Droid 1", "Tier 1", "Ready?", "Droid 2", "Tier 2", "Ready?",
                  "Droid 3", "Tier 3", "Ready?", "ALL REQUIREMENTS MET?"]
    for i, h in enumerate(rt_headers, start=1):
        rt.cell(row=1, column=i, value=h)
    style_header(rt, 1, 1, len(rt_headers))
    rt.freeze_panes = "A2"

    row = 2
    for rb in rebirths_data:
        level = rb["level"]
        credits = rb["creditsRequired"]
        reqs = [(d["name"], d["tier"]) for d in rb["requiredDroids"]]
        rt.cell(row=row, column=1, value=level).font = BOLD
        rt.cell(row=row, column=2, value=credits).font = BLUE
        rt.cell(row=row, column=2).number_format = '#,##0'
        rt.cell(row=row, column=3, value=f"=IF(Dashboard!$B$2>=B{row},\"YES\",\"NO\")")
        col = 4
        ready_cols = []
        for name, tier in reqs:
            rt.cell(row=row, column=col, value=name)
            rt.cell(row=row, column=col + 1, value=tier)
            req_rank = f'VLOOKUP({get_column_letter(col+1)}{row},Lists!$A$2:$B$7,2,FALSE)'
            owned_rank = f'IFERROR(VLOOKUP({get_column_letter(col)}{row},\'Droid Database\'!$A:$I,9,FALSE),0)'
            status_col = col + 2
            rt.cell(row=row, column=status_col, value=f'=IF({owned_rank}>={req_rank},"YES","NO")')
            ready_cols.append(get_column_letter(status_col) + str(row))
            col += 3
        overall_col = col
        rt.cell(row=row, column=overall_col,
                value=f'=IF(AND(C{row}="YES",{",".join([c+"=\"YES\"" for c in ready_cols])}),"READY","NOT YET")')
        for c in range(1, overall_col + 1):
            rt.cell(row=row, column=c).border = BORDER
            rt.cell(row=row, column=c).alignment = Alignment(horizontal="center")
        row += 1

    last_rt_row = row - 1
    green_fill = PatternFill("solid", start_color="C6EFCE", end_color="C6EFCE")
    red_fill = PatternFill("solid", start_color="FFC7CE", end_color="FFC7CE")
    overall_col_letter = get_column_letter(len(rt_headers))
    rt.conditional_formatting.add(f"{overall_col_letter}2:{overall_col_letter}{last_rt_row}",
                                   FormulaRule(formula=[f'{overall_col_letter}2="READY"'], fill=green_fill))
    rt.conditional_formatting.add(f"{overall_col_letter}2:{overall_col_letter}{last_rt_row}",
                                   FormulaRule(formula=[f'{overall_col_letter}2="NOT YET"'], fill=red_fill))
    for letter in ["C", "F", "I", "L"]:
        rt.conditional_formatting.add(f"{letter}2:{letter}{last_rt_row}",
                                       FormulaRule(formula=[f'{letter}2="YES"'], fill=green_fill))
        rt.conditional_formatting.add(f"{letter}2:{letter}{last_rt_row}",
                                       FormulaRule(formula=[f'{letter}2="NO"'], fill=red_fill))
    autosize(rt, {"A": 9, "B": 16, "C": 12, "D": 16, "E": 10, "F": 8, "G": 16, "H": 10,
                  "I": 8, "J": 16, "K": 10, "L": 8, "M": 20})

    # ---------------- Dashboard ----------------
    dash = wb.create_sheet("Dashboard")
    dash["A1"] = "Droid Tycoon Companion"
    dash["A1"].font = TITLE_FONT
    dash["A2"] = "Fortnite Star Wars: Droid Tycoon Tracker"
    dash["A2"].font = Font(name=FONT, italic=True, size=11, color="666666")

    dash["A4"] = "Your Current Credits"
    dash["A4"].font = BOLD
    dash["B4"] = 0
    dash["B4"].fill = INPUT_FILL
    dash["B4"].font = BLUE
    dash["B4"].number_format = '#,##0'
    dash["A5"] = "(enter this each time you play — used to check rebirth readiness)"
    dash["A5"].font = Font(name=FONT, italic=True, size=9, color="666666")
    dash["B2"] = "=B4"

    dash["A7"] = "Summary"
    dash["A7"].font = BOLD
    dash["A7"].fill = SUBHEADER_FILL
    dash["B7"].fill = SUBHEADER_FILL
    dash["A8"] = "Droids Owned"
    dash["B8"] = f"=COUNTIF('Droid Database'!H2:H{last_droid_row},\"<>None\")"
    dash["A9"] = "Total Droids in Database"
    dash["B9"] = last_droid_row - start_row + 1
    dash["A10"] = "Total Income / Second"
    dash["B10"] = f"=SUM('Droid Database'!J2:J{last_droid_row})"
    dash["B10"].number_format = '#,##0'
    dash["A11"] = "Total Income / Hour"
    dash["B11"] = "=B10*3600"
    dash["B11"].number_format = '#,##0'

    dash["A13"] = "Highest Rebirth Completed (enter manually)"
    dash["A13"].font = BOLD
    dash["B13"] = 0
    dash["B13"].fill = INPUT_FILL
    dash["B13"].font = BLUE

    dash["A15"] = "Next Rebirth Target"
    dash["A15"].font = BOLD
    dash["A15"].fill = SUBHEADER_FILL
    dash["B15"].fill = SUBHEADER_FILL
    dash["A16"] = "Next Rebirth Level"
    dash["B16"] = "=B13+1"
    dash["A17"] = "Credits Needed"
    dash["B17"] = "=IFERROR(INDEX('Rebirth Tracker'!B:B,MATCH(B16,'Rebirth Tracker'!A:A,0)),\"—\")"
    dash["B17"].number_format = '#,##0'
    dash["A18"] = "Credits Still Needed"
    dash["B18"] = '=IF(ISNUMBER(B17),MAX(B17-B4,0),"—")'
    dash["B18"].number_format = '#,##0'
    dash["A19"] = "Droid Requirements Met?"
    dash["B19"] = "=IFERROR(INDEX('Rebirth Tracker'!M:M,MATCH(B16,'Rebirth Tracker'!A:A,0)),\"—\")"
    dash["A20"] = "Overall Status"
    dash["B20"] = "=IFERROR(INDEX('Rebirth Tracker'!M:M,MATCH(B16,'Rebirth Tracker'!A:A,0)),\"—\")"

    dash["A22"] = "How to use this workbook"
    dash["A22"].font = BOLD
    tips = [
        "1. Go to 'Quick Entry' — pick a droid from the dropdown in column A, set Owned? to Yes, then pick its Tier.",
        "2. 'Droid Database' updates automatically to show your full collection and total income.",
        "3. Update 'Your Current Credits' above each time you check in.",
        "4. Check 'Rebirth Tracker' to see exactly which rebirth levels you're ready for (green = ready).",
        "5. 'Upgrade Costs' shows how many Upgrade Chips each rarity needs to level up a tier.",
        "6. This file is auto-generated from database/droids.json and database/rebirths.json — "
        "edit those files (not this one) to update game data, then let the GitHub Action rebuild it. "
        "Your Quick Entry progress lives only in your own downloaded copy.",
    ]
    for i, t in enumerate(tips, start=23):
        dash[f"A{i}"] = t
        dash[f"A{i}"].font = Font(name=FONT, size=10)

    autosize(dash, {"A": 42, "B": 18})
    dash.sheet_view.showGridLines = False

    desired_order = ["Dashboard", "Quick Entry", "Droid Database", "Rebirth Tracker", "Upgrade Costs", "Lists"]
    wb._sheets = [wb[name] for name in desired_order]
    wb["Lists"].sheet_state = "hidden"

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"Wrote {OUT_PATH} ({len(droids_data)} droids, {len(rebirths_data)} rebirth levels)")


if __name__ == "__main__":
    main()
